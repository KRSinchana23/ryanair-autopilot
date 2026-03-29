"""
outputs/report_generator.py
Generates two outputs:
1. Excel financial model with 3 scenarios (colour-coded, formatted)
2. PDF shareholder report

Libraries used:
- openpyxl: Excel generation (open source, no Excel required)
- reportlab: PDF generation (open source)
"""

import os
from datetime import datetime
from typing import Dict, Any

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY

from pipeline.transform import StructuredCompanyData
from pipeline.financial_model import FinancialModel

os.makedirs("outputs", exist_ok=True)


# Excel Generator

SCENARIO_COLORS = {
    "base":     {"fill": "DBEAFE", "font": "1d4ed8", "header": "2563eb"},
    "upside":   {"fill": "DCFCE7", "font": "15803d", "header": "16a34a"},
    "downside": {"fill": "FEE2E2", "font": "b91c1c", "header": "dc2626"},
}

def make_border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def header_style(ws, cell, text, hex_color, font_size=11, bold=True):
    cell.value = text
    cell.font = Font(bold=bold, color="FFFFFF", size=font_size)
    cell.fill = PatternFill("solid", fgColor=hex_color)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = make_border()

def data_cell(cell, value, fmt=None, bold=False, color=None):
    cell.value = value
    if fmt:
        cell.number_format = fmt
    cell.font = Font(bold=bold, color=color or "000000", size=10)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.border = make_border("thin")


def generate_excel(
    company_data: StructuredCompanyData,
    financial_model: FinancialModel,
    output_path: str = "outputs/ryanair_financial_model.xlsx"
) -> str:
    print("[EXCEL] Generating financial model spreadsheet")

    wb = Workbook()

    # Sheet 1: Overview 
    ws_overview = wb.active
    ws_overview.title = "Overview"
    ws_overview.column_dimensions["A"].width = 30
    ws_overview.column_dimensions["B"].width = 22

    ws_overview["A1"] = "RYANAIR HOLDINGS PLC"
    ws_overview["A1"].font = Font(bold=True, size=16, color="1d4ed8")
    ws_overview["A2"] = "Corporate Finance Autopilot — Financial Model"
    ws_overview["A2"].font = Font(size=11, color="6b7280")
    ws_overview["A3"] = f"Generated: {datetime.now().strftime('%d %B %Y, %H:%M')}"
    ws_overview["A3"].font = Font(size=9, color="9ca3af")

    ws_overview["A5"] = "⚠️ DISCLAIMER"
    ws_overview["A5"].font = Font(bold=True, color="b91c1c")
    ws_overview["A6"] = "This model is for educational purposes only."
    ws_overview["A7"] = "It is NOT investment advice. Label uncertainty."
    ws_overview["A8"] = "Sources: yfinance / Yahoo Finance (public data)."

    row = 10
    ws_overview.cell(row, 1, "KEY METRICS").font = Font(bold=True, size=11)
    row += 1

    metrics = [
        ("Company", company_data.profile.name),
        ("Ticker", company_data.profile.ticker),
        ("Exchange", company_data.profile.exchange),
        ("Sector", company_data.profile.sector),
        ("Current Price", company_data.market.current_price),
        ("Market Cap (€bn)", company_data.market.market_cap_eur_bn),
        ("52W High", company_data.market.fifty_two_week_high),
        ("52W Low", company_data.market.fifty_two_week_low),
        ("Trailing P/E", company_data.market.trailing_pe),
        ("Forward P/E", company_data.market.forward_pe),
        ("Revenue (€)", company_data.income.total_revenue),
        ("EBITDA (€)", company_data.income.ebitda),
        ("Operating Margin %", company_data.income.operating_margin_pct),
        ("Net Margin %", company_data.income.net_margin_pct),
        ("Total Debt (€)", company_data.balance.total_debt),
        ("Total Cash (€)", company_data.balance.total_cash),
        ("Net Debt (€)", company_data.balance.net_debt),
        ("Return on Equity %", company_data.balance.return_on_equity_pct),
        ("Analyst Target", company_data.market.analyst_target_price),
    ]

    for label, value in metrics:
        cell_a = ws_overview.cell(row, 1, label)
        cell_a.font = Font(bold=True, size=10)
        cell_a.border = make_border()
        cell_b = ws_overview.cell(row, 2, value if value is not None else "N/A")
        cell_b.font = Font(size=10)
        cell_b.border = make_border()
        cell_b.alignment = Alignment(horizontal="right")
        row += 1

    # Sheet 2: 3-Scenario Model
    ws_model = wb.create_sheet("Financial Model")
    ws_model.column_dimensions["A"].width = 32

    title_cell = ws_model["A1"]
    title_cell.value = "3-SCENARIO FINANCIAL MODEL"
    title_cell.font = Font(bold=True, size=14, color="1e293b")

    ws_model["A2"] = f"Base Year: {financial_model.base_year} | Company: {financial_model.company_name}"
    ws_model["A2"].font = Font(size=10, color="64748b")

    # Column layout: A=metric, B-D=base years, E-G=upside years, H-J=downside years
    col_offset = {"base": 2, "upside": 6, "downside": 10}
    years = [p.year for p in list(financial_model.scenarios["base"].projections)]

    row = 4
    # Scenario headers
    for scenario_key, start_col in col_offset.items():
        scenario = financial_model.scenarios[scenario_key]
        colors_cfg = SCENARIO_COLORS[scenario_key]
        # Merge header across 4 cols
        ws_model.merge_cells(
            start_row=row, start_column=start_col,
            end_row=row, end_column=start_col + 3
        )
        cell = ws_model.cell(row, start_col, scenario.label.upper())
        header_style(ws_model, cell, scenario.label.upper(), colors_cfg["header"], font_size=11)

    row += 1

    # Assumption sub-headers
    ws_model.cell(row, 1, "Assumptions").font = Font(bold=True, size=9, color="64748b")
    for scenario_key, start_col in col_offset.items():
        scenario = financial_model.scenarios[scenario_key]
        colors_cfg = SCENARIO_COLORS[scenario_key]
        assumptions_text = (
            f"Pax Growth: {scenario.assumptions['passenger_growth_pct']:+.1f}% | "
            f"Fuel: {scenario.assumptions['fuel_cost_delta_pct']:+.1f}% | "
            f"Yield: {scenario.assumptions['yield_change_pct']:+.1f}%"
        )
        ws_model.merge_cells(
            start_row=row, start_column=start_col,
            end_row=row, end_column=start_col + 3
        )
        c = ws_model.cell(row, start_col, assumptions_text)
        c.font = Font(size=8, color=colors_cfg["font"])
        c.fill = PatternFill("solid", fgColor=colors_cfg["fill"])
        c.alignment = Alignment(horizontal="center")

    row += 1

    # Year headers
    ws_model.cell(row, 1, "Metric").font = Font(bold=True)
    for scenario_key, start_col in col_offset.items():
        colors_cfg = SCENARIO_COLORS[scenario_key]
        for i, year in enumerate(years):
            c = ws_model.cell(row, start_col + i, str(year) + "E")
            c.font = Font(bold=True, color="FFFFFF", size=9)
            c.fill = PatternFill("solid", fgColor=colors_cfg["header"])
            c.alignment = Alignment(horizontal="center")
            c.border = make_border()
        # Summary col
        c = ws_model.cell(row, start_col + 3, "3Y Summary")
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.fill = PatternFill("solid", fgColor="475569")
        c.alignment = Alignment(horizontal="center")
        c.border = make_border()

    row += 1

    # Data rows
    metric_rows = [
        ("Passengers (m)", "passengers_m", "#,##0.0"),
        ("Revenue per Pax (€)", "revenue_per_pax_eur", "#,##0.00"),
        ("Total Revenue (€bn)", "total_revenue_eur_bn", "#,##0.000"),
        ("Fuel Cost (€bn)", "fuel_cost_eur_bn", "#,##0.000"),
        ("Other OpEx (€bn)", "other_opex_eur_bn", "#,##0.000"),
        ("EBIT (€bn)", "ebit_eur_bn", "#,##0.000"),
        ("EBIT Margin %", "ebit_margin_pct", "#,##0.0"),
    ]

    for metric_label, metric_key, fmt in metric_rows:
        ws_model.cell(row, 1, metric_label).font = Font(bold=True, size=10)
        ws_model.cell(row, 1).border = make_border()
        ws_model.cell(row, 1).column_dimensions if False else None

        for scenario_key, start_col in col_offset.items():
            scenario = financial_model.scenarios[scenario_key]
            colors_cfg = SCENARIO_COLORS[scenario_key]
            values = [getattr(p, metric_key) for p in scenario.projections]

            for i, val in enumerate(values):
                c = ws_model.cell(row, start_col + i, val)
                c.number_format = fmt
                c.font = Font(size=10, color=colors_cfg["font"])
                c.fill = PatternFill("solid", fgColor=colors_cfg["fill"])
                c.border = make_border()
                c.alignment = Alignment(horizontal="right")

            # 3Y summary - total revenue, avg for others
            if metric_key == "total_revenue_eur_bn":
                summary_val = scenario.summary["total_revenue_3y_eur_bn"]
            elif metric_key == "ebit_margin_pct":
                summary_val = scenario.summary["avg_ebit_margin_pct"]
            else:
                summary_val = sum(values) / len(values)

            c = ws_model.cell(row, start_col + 3, summary_val)
            c.number_format = fmt
            c.font = Font(bold=True, size=10, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="475569")
            c.border = make_border()
            c.alignment = Alignment(horizontal="right")

        row += 1

    # Sheet 3: Key Drivers
    ws_drivers = wb.create_sheet("Key Drivers")
    ws_drivers["A1"] = "KEY FINANCIAL DRIVERS"
    ws_drivers["A1"].font = Font(bold=True, size=13)
    row = 3
    for i, driver in enumerate(financial_model.key_drivers, 1):
        ws_drivers.cell(row, 1, f"{i}. {driver}").font = Font(size=11)
        row += 1

    row += 2
    ws_drivers.cell(row, 1, "DATA SOURCES").font = Font(bold=True)
    row += 1
    sources = [
        "yfinance: Yahoo Finance public API (stock data, financials)",
        "Ryanair Investor Relations (public annual reports)",
        "Company website: ryanair.com (public information)",
    ]
    for src in sources:
        ws_drivers.cell(row, 1, f"• {src}").font = Font(size=10)
        row += 1

    wb.save(output_path)
    print(f"[EXCEL] ✓ Saved to {output_path}")
    return output_path


# PDF Generator 

def generate_pdf(
    company_data: StructuredCompanyData,
    financial_model: FinancialModel,
    agent_results: Dict[str, Any],
    output_path: str = "outputs/ryanair_report.pdf"
) -> str:
    print("[PDF] Generating shareholder report")

    doc = SimpleDocTemplate(
        output_path,
        pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm
    )

    styles = getSampleStyleSheet()
    story = []

    # Custom styles
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Title"],
        fontSize=24, textColor=colors.HexColor("#1d4ed8"),
        spaceAfter=6, fontName="Helvetica-Bold"
    )
    subtitle_style = ParagraphStyle(
        "Subtitle",
        parent=styles["Normal"],
        fontSize=12, textColor=colors.HexColor("#6b7280"),
        spaceAfter=4
    )
    h1_style = ParagraphStyle(
        "H1Custom",
        parent=styles["Heading1"],
        fontSize=14, textColor=colors.HexColor("#1e293b"),
        spaceBefore=14, spaceAfter=6, fontName="Helvetica-Bold"
    )
    body_style = ParagraphStyle(
        "BodyCustom",
        parent=styles["Normal"],
        fontSize=10, leading=14,
        textColor=colors.HexColor("#374151"),
        spaceAfter=8, alignment=TA_JUSTIFY
    )
    warning_style = ParagraphStyle(
        "Warning",
        parent=styles["Normal"],
        fontSize=9, textColor=colors.HexColor("#b91c1c"),
        backColor=colors.HexColor("#FEF2F2"),
        borderPadding=6, spaceAfter=12
    )

    # Title Page 
    story.append(Spacer(1, 1*cm))
    story.append(Paragraph("RYANAIR HOLDINGS PLC", title_style))
    story.append(Paragraph("Corporate Finance Autopilot — Equity Research Report", subtitle_style))
    story.append(Paragraph(f"Ticker: {company_data.profile.ticker} | {datetime.now().strftime('%B %Y')}", subtitle_style))
    story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor("#2563eb")))
    story.append(Spacer(1, 0.3*cm))

    story.append(Paragraph(
        "DISCLAIMER: This report is produced by an automated AI system for educational purposes only. "
        "It is NOT investment advice. All figures should be independently verified. "
        "Sources: Yahoo Finance (yfinance), public company filings.",
        warning_style
    ))

    # Key Metrics Table 
    story.append(Paragraph("KEY METRICS SNAPSHOT", h1_style))

    market = company_data.market
    income = company_data.income

    metrics_data = [
        ["Metric", "Value", "Metric", "Value"],
        ["Current Price", str(market.current_price) if market.current_price else "N/A",
         "Market Cap (€bn)", str(market.market_cap_eur_bn) if market.market_cap_eur_bn else "N/A"],
        ["52W High", str(market.fifty_two_week_high) if market.fifty_two_week_high else "N/A",
         "52W Low", str(market.fifty_two_week_low) if market.fifty_two_week_low else "N/A"],
        ["Trailing P/E", str(market.trailing_pe) if market.trailing_pe else "N/A",
         "Forward P/E", str(market.forward_pe) if market.forward_pe else "N/A"],
        ["Op. Margin %", str(income.operating_margin_pct) if income.operating_margin_pct else "N/A",
         "Net Margin %", str(income.net_margin_pct) if income.net_margin_pct else "N/A"],
        ["Revenue Growth %", str(income.revenue_growth_pct) if income.revenue_growth_pct else "N/A",
         "Analyst Target", str(market.analyst_target_price) if market.analyst_target_price else "N/A"],
    ]

    metrics_table = Table(metrics_data, colWidths=[4*cm, 3.5*cm, 4*cm, 3.5*cm])
    metrics_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563eb")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F8FAFC")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#F8FAFC"), colors.white]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("ALIGN", (3, 0), (3, -1), "RIGHT"),
        ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 1), (2, -1), "Helvetica-Bold"),
        ("PADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(metrics_table)
    story.append(Spacer(1, 0.4*cm))

    # AI-Generated Sections
    sections = agent_results.get("sections", {})
    section_order = [
        ("executive_summary", "EXECUTIVE SUMMARY"),
        ("business_overview", "BUSINESS OVERVIEW"),
        ("financial_analysis", "FINANCIAL ANALYSIS"),
        ("risk_factors", "RISK FACTORS"),
        ("investment_thesis", "INVESTMENT THESIS"),
        ("strategic_options", "STRATEGIC OPTIONS & FUNDING"),
    ]

    for section_key, section_title in section_order:
        content = sections.get(section_key)
        if content:
            story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#CBD5E1")))
            story.append(Paragraph(section_title, h1_style))
            story.append(Paragraph(content.replace("\n", "<br/>"), body_style))

    # Scenario Summary Table 
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#CBD5E1")))
    story.append(Paragraph("3-SCENARIO FINANCIAL PROJECTIONS", h1_style))

    scenario_table_data = [["Scenario", "3Y Revenue (€bn)", "Avg EBIT Margin %", "Final Year Rev (€bn)"]]
    scenario_colors_map = {
        "base": colors.HexColor("#DBEAFE"),
        "upside": colors.HexColor("#DCFCE7"),
        "downside": colors.HexColor("#FEE2E2"),
    }

    for key, scenario in financial_model.scenarios.items():
        scenario_table_data.append([
            scenario.label,
            f"€{scenario.summary['total_revenue_3y_eur_bn']:.2f}bn",
            f"{scenario.summary['avg_ebit_margin_pct']:.1f}%",
            f"€{scenario.summary['final_year_revenue_eur_bn']:.3f}bn",
        ])

    scenario_table = Table(scenario_table_data, colWidths=[4*cm, 4*cm, 4*cm, 4*cm])
    scenario_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e293b")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 1), (-1, 1), scenario_colors_map["base"]),
        ("BACKGROUND", (0, 2), (-1, 2), scenario_colors_map["upside"]),
        ("BACKGROUND", (0, 3), (-1, 3), scenario_colors_map["downside"]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("PADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(scenario_table)

    # Agent Trace Summary 
    story.append(Spacer(1, 0.5*cm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#CBD5E1")))
    story.append(Paragraph("AI AGENT TRACE (Observable Steps)", h1_style))
    story.append(Paragraph(
        f"This report was generated by a multi-step AI agent. "
        f"Total steps executed: {agent_results.get('steps_count', 0)}. "
        f"Each step's tool calls are logged in the API response for full observability.",
        body_style
    ))

    doc.build(story)
    print(f"[PDF] Saved to {output_path}")
    return output_path
